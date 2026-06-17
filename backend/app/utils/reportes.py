"""Helpers genéricos para exportar reportes tabulares en PDF y Excel.

Cada reporte se describe como una lista de secciones:
    [{'titulo': str | None, 'columnas': [...], 'filas': [[...], ...]}]
"""
import io
from flask import Response
from .timezone import ahora_bolivia

MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _ascii(texto):
    """fpdf2 con fuentes core solo soporta latin-1; normalizamos acentos."""
    return ('' if texto is None else str(texto)).encode('latin-1', 'replace').decode('latin-1')


def _formatear(valor):
    if isinstance(valor, float):
        return f'{valor:,.2f}'
    return '' if valor is None else str(valor)


def exportar_secciones_xlsx(titulo, subtitulos, secciones, nombre_archivo):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    relleno = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    negrita_blanca = Font(bold=True, color='FFFFFF')

    ws.append([titulo])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    for sub in subtitulos:
        ws.append([sub])
    ws.append([f'Generado: {ahora_bolivia().strftime("%Y-%m-%d %H:%M")}'])
    ws.append([])

    max_cols = 1
    for seccion in secciones:
        if seccion.get('titulo'):
            ws.append([seccion['titulo']])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append(seccion['columnas'])
        for celda in ws[ws.max_row]:
            if celda.value is not None:
                celda.font = negrita_blanca
                celda.fill = relleno
        for fila in seccion['filas']:
            ws.append(list(fila))
        ws.append([])
        max_cols = max(max_cols, len(seccion['columnas']))

    for i in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 32 if i == 1 else 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype=MIME_XLSX,
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}.xlsx'},
    )


def exportar_secciones_pdf(titulo, subtitulos, secciones, nombre_archivo, orientacion='P'):
    from fpdf import FPDF

    pdf = FPDF(orientation=orientacion, unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 15)
    pdf.cell(0, 9, _ascii(titulo), ln=1)
    pdf.set_font('Helvetica', '', 10)
    for sub in subtitulos:
        pdf.cell(0, 6, _ascii(sub), ln=1)
    pdf.cell(0, 6, f'Generado: {ahora_bolivia().strftime("%Y-%m-%d %H:%M")}', ln=1)
    pdf.ln(3)

    ancho_util = pdf.w - pdf.l_margin - pdf.r_margin
    for seccion in secciones:
        columnas = seccion['columnas']
        n = len(columnas)
        if n > 1:
            base = ancho_util / (n + 0.6)
            anchos = [base * 1.6] + [base] * (n - 1)
        else:
            anchos = [ancho_util]

        if seccion.get('titulo'):
            pdf.set_font('Helvetica', 'B', 11)
            pdf.cell(0, 7, _ascii(seccion['titulo']), ln=1)

        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(31, 78, 120)
        pdf.set_text_color(255, 255, 255)
        for ancho, col in zip(anchos, columnas):
            pdf.cell(ancho, 6, _ascii(col), border=1, fill=True)
        pdf.ln()

        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 8)
        for fila in seccion['filas']:
            for ancho, valor in zip(anchos, fila):
                texto = _ascii(_formatear(valor))
                limite = max(4, int(ancho / 1.7))
                if len(texto) > limite:
                    texto = texto[:limite - 1] + '.'
                align = 'R' if isinstance(valor, (int, float)) else 'L'
                pdf.cell(ancho, 5.5, texto, border=1, align=align)
            pdf.ln()
        pdf.ln(3)

    salida = pdf.output(dest='S')
    if isinstance(salida, str):
        salida = salida.encode('latin-1')
    return Response(
        bytes(salida),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}.pdf'},
    )
