import csv
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, Response
from flask_jwt_extended import jwt_required
from sqlalchemy import func
from ...extensions import db
from ...models.auditoria.auditoria import Bitacora, BitacoraDetalle
from ...models.seguridad.auth import Usuario
from ...utils.permisos import requiere_permiso
from ...utils.timezone import ahora_bolivia

bp = Blueprint('auditoria', __name__)


def _aplicar_filtros(consulta, args):
    """Aplica los filtros comunes (compartidos por listar y exportar)."""
    q              = args.get('q', '').strip()
    usuario_filtro = args.get('usuario', '').strip()
    id_usuario     = args.get('id_usuario', type=int)
    accion_filtro  = args.get('accion', '').strip()
    modulo_filtro  = args.get('modulo', '').strip()
    fecha_desde    = args.get('fecha_desde', '').strip()
    fecha_hasta    = args.get('fecha_hasta', '').strip()

    if q:
        consulta = consulta.filter(Bitacora.descripcion.ilike(f'%{q}%'))
    if usuario_filtro:
        consulta = consulta.filter(Usuario.username.ilike(f'%{usuario_filtro}%'))
    if id_usuario:
        consulta = consulta.filter(Bitacora.id_usuario == id_usuario)
    if accion_filtro:
        consulta = consulta.filter(Bitacora.accion == accion_filtro)
    if modulo_filtro:
        consulta = consulta.filter(Bitacora.modulo == modulo_filtro)
    if fecha_desde:
        consulta = consulta.filter(Bitacora.fecha >= fecha_desde)
    if fecha_hasta:
        consulta = consulta.filter(Bitacora.fecha <= f'{fecha_hasta} 23:59:59')
    return consulta


@bp.get('/')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def listar():
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(200, max(1, int(request.args.get('per_page', 50))))

    consulta = (
        db.session.query(Bitacora, Usuario.username)
        .outerjoin(Usuario, Bitacora.id_usuario == Usuario.id)
        .order_by(Bitacora.fecha.desc())
    )
    consulta = _aplicar_filtros(consulta, request.args)

    total = consulta.count()
    filas = consulta.offset((page - 1) * per_page).limit(per_page).all()

    items = [_serializar(b, username) for b, username in filas]

    return jsonify({
        'items': items,
        'total': total,
        'page': page,
        'pages': max(1, (total + per_page - 1) // per_page),
        'per_page': per_page,
    })


@bp.get('/<int:id_bitacora>')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def obtener(id_bitacora):
    b = db.get_or_404(Bitacora, id_bitacora)
    username = b.usuario.username if b.usuario else None
    data = _serializar(b, username)
    data['detalles'] = [
        {
            'campo': d.campo,
            'valor_anterior': d.valor_anterior,
            'valor_nuevo': d.valor_nuevo,
        }
        for d in b.detalles
    ]
    return jsonify(data)


@bp.get('/acciones')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def listar_acciones():
    acciones = db.session.query(Bitacora.accion).distinct().order_by(Bitacora.accion).all()
    return jsonify([a[0] for a in acciones])


@bp.get('/modulos')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def listar_modulos():
    modulos = (db.session.query(Bitacora.modulo)
               .filter(Bitacora.modulo.isnot(None))
               .distinct()
               .order_by(Bitacora.modulo)
               .all())
    return jsonify([m[0] for m in modulos])


# ── CU48: Consultar auditoría por usuario ───────────────────────

@bp.get('/usuarios')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def listar_usuarios_con_actividad():
    """Usuarios que tienen al menos una acción registrada en bitácora."""
    filas = (
        db.session.query(Usuario.id, Usuario.username, func.count(Bitacora.id))
        .join(Bitacora, Bitacora.id_usuario == Usuario.id)
        .group_by(Usuario.id, Usuario.username)
        .order_by(Usuario.username)
        .all()
    )
    return jsonify([
        {'id': uid, 'username': uname, 'acciones': total}
        for uid, uname, total in filas
    ])


# ── CU47: Exportar log de auditoría (CSV / Excel / PDF) ─────────

@bp.get('/exportar')
@jwt_required()
@requiere_permiso('gestionar_usuarios')
def exportar():
    formato = (request.args.get('formato') or 'csv').lower()
    if formato not in ('csv', 'pdf', 'xlsx'):
        return jsonify({'error': 'Formato no soportado (use csv, xlsx o pdf)'}), 400

    consulta = (
        db.session.query(Bitacora, Usuario.username)
        .outerjoin(Usuario, Bitacora.id_usuario == Usuario.id)
        .order_by(Bitacora.fecha.desc())
    )
    consulta = _aplicar_filtros(consulta, request.args)
    filas = consulta.limit(10000).all()

    # E1: sin registros para los filtros aplicados
    if not filas:
        return jsonify({'error': 'No hay registros para los filtros aplicados'}), 404

    marca = ahora_bolivia().strftime('%Y%m%d_%H%M')
    if formato == 'csv':
        return _exportar_csv(filas, marca)
    if formato == 'xlsx':
        return _exportar_xlsx(filas, marca)
    return _exportar_pdf(filas, marca)


def _exportar_csv(filas, marca):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Fecha', 'Usuario', 'Módulo', 'Acción', 'Descripción', 'IP'])
    for b, username in filas:
        writer.writerow([
            b.fecha.strftime('%Y-%m-%d %H:%M:%S') if b.fecha else '',
            username or 'sistema',
            b.modulo or '',
            b.accion or '',
            (b.descripcion or '').replace('\n', ' '),
            b.ip or '',
        ])
    contenido = '﻿' + buffer.getvalue()  # BOM para acentos en Excel
    return Response(
        contenido,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=auditoria_{marca}.csv'},
    )


def _exportar_xlsx(filas, marca):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Auditoría'

    encabezados = ['Fecha', 'Usuario', 'Módulo', 'Acción', 'Descripción', 'IP']
    ws.append(encabezados)
    relleno = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = relleno

    for b, username in filas:
        ws.append([
            b.fecha.strftime('%Y-%m-%d %H:%M:%S') if b.fecha else '',
            username or 'sistema',
            b.modulo or '',
            b.accion or '',
            (b.descripcion or '').replace('\n', ' '),
            b.ip or '',
        ])

    anchos = [20, 18, 16, 28, 70, 16]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=auditoria_{marca}.xlsx'},
    )


def _exportar_pdf(filas, marca):
    from fpdf import FPDF

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, 'ServiControl - Log de auditoria', ln=1)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 5, f'Generado: {ahora_bolivia().strftime("%Y-%m-%d %H:%M")}  -  {len(filas)} registro(s)', ln=1)
    pdf.ln(2)

    # Anchos de columna (total ~ 273mm en A4 horizontal con márgenes)
    cols = [(34, 'Fecha'), (28, 'Usuario'), (26, 'Modulo'), (45, 'Accion'), (110, 'Descripcion'), (30, 'IP')]
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(230, 230, 230)
    for ancho, titulo in cols:
        pdf.cell(ancho, 6, titulo, border=1, fill=True)
    pdf.ln()

    pdf.set_font('Helvetica', '', 7)
    for b, username in filas:
        valores = [
            b.fecha.strftime('%Y-%m-%d %H:%M') if b.fecha else '',
            username or 'sistema',
            b.modulo or '',
            b.accion or '',
            (b.descripcion or '').replace('\n', ' '),
            b.ip or '',
        ]
        for (ancho, _), valor in zip(cols, valores):
            texto = _ascii(valor)
            # Truncar para que entre en la celda
            limite = max(4, int(ancho / 1.5))
            if len(texto) > limite:
                texto = texto[:limite - 1] + '.'
            pdf.cell(ancho, 5, texto, border=1)
        pdf.ln()

    salida = pdf.output(dest='S')
    if isinstance(salida, str):
        salida = salida.encode('latin-1')
    return Response(
        bytes(salida),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=auditoria_{marca}.pdf'},
    )


def _ascii(texto):
    """fpdf2 con fuentes core solo soporta latin-1; normalizamos acentos."""
    return (texto or '').encode('latin-1', 'replace').decode('latin-1')


# ── CU49: Generar reporte de actividad ──────────────────────────

def _rango_reporte(args):
    """Lee y valida el rango de fechas; por defecto últimos 30 días."""
    hoy = ahora_bolivia().date()
    desde_raw = (args.get('desde') or '').strip()
    hasta_raw = (args.get('hasta') or '').strip()

    desde = datetime.strptime(desde_raw, '%Y-%m-%d').date() if desde_raw else hoy - timedelta(days=30)
    hasta = datetime.strptime(hasta_raw, '%Y-%m-%d').date() if hasta_raw else hoy
    return desde, hasta


def _construir_reporte(desde, hasta, id_usuario=None):
    """Métricas + desgloses del período. id_usuario filtra toda la actividad."""
    from ...models.proyectos.proyecto import Proyecto, EstadoProyecto
    from ...models.ordenes.orden import OrdenTrabajo, EstadoOrden
    from ...models.mantenimiento.mantenimiento import Mantenimiento
    from ...models.finanzas.finanzas import Pago

    fin = f'{hasta.isoformat()} 23:59:59'

    def filtro_usuario(consulta, columna):
        return consulta.filter(columna == id_usuario) if id_usuario else consulta

    # ── Métricas principales ──
    proy_q = Proyecto.query.filter(Proyecto.fecha_creacion >= desde, Proyecto.fecha_creacion <= fin)
    proy_q = filtro_usuario(proy_q, Proyecto.id_usuario)
    proyectos_creados = proy_q.count()

    ot_q = OrdenTrabajo.query.filter(OrdenTrabajo.fecha_creacion >= desde, OrdenTrabajo.fecha_creacion <= fin)
    ot_q = filtro_usuario(ot_q, OrdenTrabajo.id_usuario)
    ot_generadas = ot_q.count()

    mant_q = Mantenimiento.query.filter(
        Mantenimiento.fecha_programada >= desde,
        Mantenimiento.fecha_programada <= hasta,
    )
    mant_q = filtro_usuario(mant_q, Mantenimiento.id_usuario)
    mantenimientos_ejecutados = mant_q.filter(Mantenimiento.estado == 'completado').count()

    pagos_q = Pago.query.filter(Pago.fecha_pago >= desde, Pago.fecha_pago <= hasta)
    pagos_q = filtro_usuario(pagos_q, Pago.id_usuario)
    pagos_registrados = pagos_q.count()
    monto_q = db.session.query(func.coalesce(func.sum(Pago.monto), 0)).filter(
        Pago.fecha_pago >= desde, Pago.fecha_pago <= hasta)
    monto_q = filtro_usuario(monto_q, Pago.id_usuario)
    monto_cobrado = float(monto_q.scalar() or 0)

    bit_q = db.session.query(Bitacora).filter(Bitacora.fecha >= desde, Bitacora.fecha <= fin)
    bit_q = filtro_usuario(bit_q, Bitacora.id_usuario)
    acciones_registradas = bit_q.count()

    usuarios_q = (
        db.session.query(func.count(func.distinct(Bitacora.id_usuario)))
        .filter(Bitacora.fecha >= desde, Bitacora.fecha <= fin,
                Bitacora.id_usuario.isnot(None))
    )
    usuarios_q = filtro_usuario(usuarios_q, Bitacora.id_usuario)
    usuarios_activos = usuarios_q.scalar() or 0

    # ── Desgloses ──
    def desglose(consulta):
        return [{'nombre': nombre or '(sin dato)', 'cantidad': int(cant)} for nombre, cant in consulta]

    proy_estado_q = (
        db.session.query(EstadoProyecto.nombre, func.count(Proyecto.id))
        .join(Proyecto, Proyecto.id_estado_proyecto == EstadoProyecto.id)
        .filter(Proyecto.fecha_creacion >= desde, Proyecto.fecha_creacion <= fin)
    )
    proy_estado_q = filtro_usuario(proy_estado_q, Proyecto.id_usuario)
    proyectos_por_estado = desglose(proy_estado_q.group_by(EstadoProyecto.nombre)
                                    .order_by(func.count(Proyecto.id).desc()).all())

    ot_estado_q = (
        db.session.query(EstadoOrden.nombre, func.count(OrdenTrabajo.id))
        .join(OrdenTrabajo, OrdenTrabajo.id_estado_orden == EstadoOrden.id)
        .filter(OrdenTrabajo.fecha_creacion >= desde, OrdenTrabajo.fecha_creacion <= fin)
    )
    ot_estado_q = filtro_usuario(ot_estado_q, OrdenTrabajo.id_usuario)
    ordenes_por_estado = desglose(ot_estado_q.group_by(EstadoOrden.nombre)
                                  .order_by(func.count(OrdenTrabajo.id).desc()).all())

    mant_estado_q = (
        db.session.query(Mantenimiento.estado, func.count(Mantenimiento.id))
        .filter(Mantenimiento.fecha_programada >= desde, Mantenimiento.fecha_programada <= hasta)
    )
    mant_estado_q = filtro_usuario(mant_estado_q, Mantenimiento.id_usuario)
    mantenimientos_por_estado = desglose(mant_estado_q.group_by(Mantenimiento.estado)
                                         .order_by(func.count(Mantenimiento.id).desc()).all())

    mant_tipo_q = (
        db.session.query(Mantenimiento.tipo, func.count(Mantenimiento.id))
        .filter(Mantenimiento.fecha_programada >= desde, Mantenimiento.fecha_programada <= hasta)
    )
    mant_tipo_q = filtro_usuario(mant_tipo_q, Mantenimiento.id_usuario)
    mantenimientos_por_tipo = desglose(mant_tipo_q.group_by(Mantenimiento.tipo).all())

    pagos_metodo_q = (
        db.session.query(Pago.metodo, func.count(Pago.id), func.coalesce(func.sum(Pago.monto), 0))
        .filter(Pago.fecha_pago >= desde, Pago.fecha_pago <= hasta)
    )
    pagos_metodo_q = filtro_usuario(pagos_metodo_q, Pago.id_usuario)
    pagos_por_metodo = [
        {'nombre': metodo or '(sin dato)', 'cantidad': int(cant), 'monto': float(monto)}
        for metodo, cant, monto in pagos_metodo_q.group_by(Pago.metodo)
                                                 .order_by(func.count(Pago.id).desc()).all()
    ]

    modulo_q = (
        db.session.query(Bitacora.modulo, func.count(Bitacora.id))
        .filter(Bitacora.fecha >= desde, Bitacora.fecha <= fin)
    )
    modulo_q = filtro_usuario(modulo_q, Bitacora.id_usuario)
    acciones_por_modulo = desglose(modulo_q.group_by(Bitacora.modulo)
                                   .order_by(func.count(Bitacora.id).desc()).limit(10).all())

    top_q = (
        db.session.query(Usuario.username, func.count(Bitacora.id))
        .join(Bitacora, Bitacora.id_usuario == Usuario.id)
        .filter(Bitacora.fecha >= desde, Bitacora.fecha <= fin)
    )
    top_q = filtro_usuario(top_q, Bitacora.id_usuario)
    top_usuarios = desglose(top_q.group_by(Usuario.username)
                            .order_by(func.count(Bitacora.id).desc()).limit(5).all())

    usuario_nombre = None
    if id_usuario:
        u = db.session.get(Usuario, id_usuario)
        usuario_nombre = u.username if u else None

    return {
        'desde': desde.isoformat(),
        'hasta': hasta.isoformat(),
        'usuario': usuario_nombre,
        'metricas': {
            'proyectos_creados':         proyectos_creados,
            'ot_generadas':              ot_generadas,
            'mantenimientos_ejecutados': mantenimientos_ejecutados,
            'pagos_registrados':         pagos_registrados,
            'monto_cobrado':             monto_cobrado,
            'usuarios_activos':          int(usuarios_activos),
            'acciones_registradas':      acciones_registradas,
        },
        'desglose': {
            'proyectos_por_estado':      proyectos_por_estado,
            'ordenes_por_estado':        ordenes_por_estado,
            'mantenimientos_por_estado': mantenimientos_por_estado,
            'mantenimientos_por_tipo':   mantenimientos_por_tipo,
            'pagos_por_metodo':          pagos_por_metodo,
            'acciones_por_modulo':       acciones_por_modulo,
            'top_usuarios':              top_usuarios,
        },
    }


@bp.get('/reporte-actividad')
@jwt_required()
@requiere_permiso('ver_reportes')
def reporte_actividad():
    try:
        desde, hasta = _rango_reporte(request.args)
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido (use AAAA-MM-DD)'}), 400

    # E2: rango inválido
    if desde > hasta:
        return jsonify({'error': 'La fecha inicial no puede ser posterior a la final'}), 422

    id_usuario = request.args.get('id_usuario', type=int)
    return jsonify(_construir_reporte(desde, hasta, id_usuario))


@bp.get('/reporte-actividad/exportar')
@jwt_required()
@requiere_permiso('ver_reportes')
def exportar_reporte_actividad():
    formato = (request.args.get('formato') or 'pdf').lower()
    if formato not in ('pdf', 'xlsx'):
        return jsonify({'error': 'Formato no soportado (use pdf o xlsx)'}), 400

    try:
        desde, hasta = _rango_reporte(request.args)
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido (use AAAA-MM-DD)'}), 400
    if desde > hasta:
        return jsonify({'error': 'La fecha inicial no puede ser posterior a la final'}), 422

    id_usuario = request.args.get('id_usuario', type=int)
    reporte = _construir_reporte(desde, hasta, id_usuario)
    marca = ahora_bolivia().strftime('%Y%m%d_%H%M')

    if formato == 'xlsx':
        return _reporte_actividad_xlsx(reporte, marca)
    return _reporte_actividad_pdf(reporte, marca)


_ETIQUETAS_METRICAS = [
    ('proyectos_creados',         'Proyectos creados'),
    ('ot_generadas',              'Órdenes de trabajo generadas'),
    ('mantenimientos_ejecutados', 'Mantenimientos ejecutados'),
    ('pagos_registrados',         'Pagos registrados'),
    ('monto_cobrado',             'Monto cobrado (Bs)'),
    ('usuarios_activos',          'Usuarios activos'),
    ('acciones_registradas',      'Acciones registradas'),
]

_ETIQUETAS_DESGLOSE = [
    ('proyectos_por_estado',      'Proyectos por estado'),
    ('ordenes_por_estado',        'Órdenes por estado'),
    ('mantenimientos_por_estado', 'Mantenimientos por estado'),
    ('mantenimientos_por_tipo',   'Mantenimientos por tipo'),
    ('pagos_por_metodo',          'Pagos por método'),
    ('acciones_por_modulo',       'Actividad por módulo'),
    ('top_usuarios',              'Usuarios más activos'),
]


def _reporte_actividad_xlsx(reporte, marca):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    relleno = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    negrita_blanca = Font(bold=True, color='FFFFFF')

    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['ServiControl — Reporte de actividad'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Período: {reporte["desde"]} al {reporte["hasta"]}'])
    if reporte.get('usuario'):
        ws.append([f'Usuario: {reporte["usuario"]}'])
    ws.append([f'Generado: {ahora_bolivia().strftime("%Y-%m-%d %H:%M")}'])
    ws.append([])
    ws.append(['Métrica', 'Valor'])
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = negrita_blanca
        celda.fill = relleno
    for clave, etiqueta in _ETIQUETAS_METRICAS:
        ws.append([etiqueta, reporte['metricas'][clave]])
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16

    ws2 = wb.create_sheet('Desglose')
    fila = 1
    for clave, titulo in _ETIQUETAS_DESGLOSE:
        items = reporte['desglose'][clave]
        if not items:
            continue
        con_monto = clave == 'pagos_por_metodo'
        ws2.cell(row=fila, column=1, value=titulo).font = Font(bold=True, size=12)
        fila += 1
        encabezados = ['Detalle', 'Cantidad'] + (['Monto (Bs)'] if con_monto else [])
        for col, texto in enumerate(encabezados, start=1):
            celda = ws2.cell(row=fila, column=col, value=texto)
            celda.font = negrita_blanca
            celda.fill = relleno
        fila += 1
        for item in items:
            ws2.cell(row=fila, column=1, value=item['nombre'])
            ws2.cell(row=fila, column=2, value=item['cantidad'])
            if con_monto:
                ws2.cell(row=fila, column=3, value=item['monto'])
            fila += 1
        fila += 1
    for i, ancho in enumerate([34, 14, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=reporte_actividad_{marca}.xlsx'},
    )


def _reporte_actividad_pdf(reporte, marca):
    from fpdf import FPDF

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 15)
    pdf.cell(0, 9, 'ServiControl - Reporte de actividad', ln=1)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, _ascii(f'Período: {reporte["desde"]} al {reporte["hasta"]}'), ln=1)
    if reporte.get('usuario'):
        pdf.cell(0, 6, _ascii(f'Usuario: {reporte["usuario"]}'), ln=1)
    pdf.cell(0, 6, f'Generado: {ahora_bolivia().strftime("%Y-%m-%d %H:%M")}', ln=1)
    pdf.ln(3)

    # Tabla de métricas principales
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 7, 'Resumen del período', ln=1)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(31, 78, 120)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(110, 7, 'Métrica'.encode('latin-1', 'replace').decode('latin-1'), border=1, fill=True)
    pdf.cell(40, 7, 'Valor', border=1, fill=True, ln=1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 9)
    for clave, etiqueta in _ETIQUETAS_METRICAS:
        valor = reporte['metricas'][clave]
        if clave == 'monto_cobrado':
            valor = f'{valor:,.2f}'
        pdf.cell(110, 6, _ascii(etiqueta), border=1)
        pdf.cell(40, 6, _ascii(str(valor)), border=1, align='R', ln=1)
    pdf.ln(4)

    # Desgloses
    for clave, titulo in _ETIQUETAS_DESGLOSE:
        items = reporte['desglose'][clave]
        if not items:
            continue
        con_monto = clave == 'pagos_por_metodo'
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, _ascii(titulo), ln=1)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(31, 78, 120)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(90, 6, 'Detalle', border=1, fill=True)
        pdf.cell(30, 6, 'Cantidad', border=1, fill=True, ln=0 if con_monto else 1)
        if con_monto:
            pdf.cell(35, 6, 'Monto (Bs)', border=1, fill=True, ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 9)
        for item in items:
            pdf.cell(90, 6, _ascii(item['nombre'])[:55], border=1)
            pdf.cell(30, 6, str(item['cantidad']), border=1, align='R', ln=0 if con_monto else 1)
            if con_monto:
                pdf.cell(35, 6, f"{item['monto']:,.2f}", border=1, align='R', ln=1)
        pdf.ln(3)

    salida = pdf.output(dest='S')
    if isinstance(salida, str):
        salida = salida.encode('latin-1')
    return Response(
        bytes(salida),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=reporte_actividad_{marca}.pdf'},
    )


def _serializar(b: Bitacora, username) -> dict:
    return {
        'id': b.id,
        'id_usuario': b.id_usuario,
        'usuario': username,
        'accion': b.accion,
        'modulo': b.modulo,
        'descripcion': b.descripcion,
        'ip': b.ip,
        'fecha': b.fecha.isoformat() if b.fecha else None,
    }
